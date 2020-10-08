import openpyxl as reader
import datetime
import random
import lxml

crashing = []

def edit_file():
    file = "Level3"
    time_index = "D"
    lat_index = "E"

    filename = file + ".xlsx"
    file = filename
    wb = reader.load_workbook((file))
    ws = wb.active
    count = 2

    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        i = 0
        time = datetime.datetime(2020, 10, 7, 14, 16, 45, 21356)
        isCrashing = False
        crash_count = 0
        while i < 59:
            if isCrashing == False:
                if random.randrange(0,100) < 5:
                    isCrashing = True


            lat = ws["A" + str(count)].value
            lon = ws["B" + str(count)].value
            footprint = ws["C" + str(count)].value

            ws["A" + str(count+(i*15371))] = lat
            ws["B" + str(count + (i * 15371))] = lon
            ws["C" + str(count + (i * 15371))] = footprint
            ws[time_index + str(count+(i*15371))] = time
            ws[time_index + str(count+(i*15371))].number_format= 'YYYY-MM-DD HH:MM:SS.00'
            ws["F" + str(count + (i * 15371))] = count - 2

            if isCrashing == False:
                ws[lat_index + str(count+(i*15371))] = (random.randrange(10,30))

            else:
                if crash_count == 0:
                    ws[lat_index + str(count + (i * 15371))] = (random.randrange(40, 60))
                    crash_count+=1

                elif crash_count == 1:
                    ws[lat_index + str(count + (i * 15371))] = (random.randrange(100, 200))
                    crash_count += 1

                elif crash_count == 2:
                    ws[lat_index + str(count + (i * 15371))] = (random.randrange(150, 400))
                    crash_count += 1

                elif crash_count == 3:
                    ws[lat_index + str(count + (i * 15371))] = (random.randrange(450, 700))
                    crash_count += 1


                elif crash_count == 4:
                    ws[lat_index + str(count + (i * 15371))] = (random.randrange(700, 1000))
                    crash_count += 1

                elif crash_count == 5:
                    ws[lat_index + str(count + (i * 15371))] = -1
                    crash_count = 0
                    isCrashing = False



            time = time + datetime.timedelta(0, 10)
            i = i + 1
        count = count + 1


    wb.save("Editted" + filename)

if __name__ == '__main__':
    edit_file()


